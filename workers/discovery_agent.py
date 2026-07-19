"""
Influencer Discovery Agent
- Accepts a search brief (market, platform, niche, follower range, count)
- Scrapes profiles via BrightData Datasets API
- Scores each profile with Claude Haiku
- Deduplicates against existing spreadsheet
- Appends new prospects to Simify_Influencer_Prospects.xlsx
"""

import os
import re
import json
import time
import argparse
from datetime import datetime
from pathlib import Path
from typing import Optional

import anthropic
import openpyxl
import requests
from dotenv import load_dotenv

load_dotenv(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"), override=True)

SPREADSHEET_PATH = Path(__file__).parent / "Simify_Influencer_Prospects.xlsx"
SPREADSHEET_COLUMNS = [
    "Channel Name", "Subscribers", "Country", "Niche", "Contact Email",
    "Management", "YouTube URL", "Est. Rate (USD)", "Status", "Notes",
]

# Maps a spreadsheet header (case/space-insensitive) to the canonical field name
# used on prospect dicts. Header-driven so append/dedup follow the sheet's actual
# column order instead of a hardcoded position — the real sheet's columns don't
# match this module's original hardcoded order, and writing positionally silently
# scrambled data into the wrong columns.
HEADER_TO_FIELD = {
    "channel name": "name", "name": "name",
    "handle": "handle",
    "subscribers": "followers", "followers": "followers",
    "platform": "platform",
    "niche": "niche",
    "country": "country",
    "contact email": "email", "email": "email",
    "management": "management",
    "youtube url": "url", "url": "url",
    "est. rate (usd)": "rate", "rate": "rate",
    "status": "status",
    "notes": "notes",
    "last contacted": "last_contacted",
}


def _sheet_fields(ws) -> list:
    """Canonical field name for each column in the sheet, in the sheet's own order."""
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    return [HEADER_TO_FIELD.get(str(h).strip().lower()) if h else None for h in header_row]


def _dedup_key(entry: dict) -> str:
    return str(entry.get("url") or entry.get("handle") or entry.get("name") or "").strip().lower()

# BrightData dataset IDs for each platform
BRIGHTDATA_DATASET_IDS = {
    "instagram": "gd_l1vikfnt1wgvvqz95w",
    "tiktok":    "gd_l7q7dkf244hwjntr0",
    "youtube":   "gd_lk538t2k2p1k3oos71",
}

BRIGHTDATA_BASE_URL = "https://api.brightdata.com/datasets/v3"
BRIGHTDATA_POLL_INTERVAL = 5   # seconds between status checks
BRIGHTDATA_POLL_TIMEOUT  = 120 # seconds before giving up


# ---------------------------------------------------------------------------
# BrightData helpers
# ---------------------------------------------------------------------------

def _brightdata_headers() -> dict:
    token = os.environ.get("BRIGHTDATA_API_TOKEN")
    if not token:
        raise EnvironmentError("BRIGHTDATA_API_TOKEN is not set in .env")
    return {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}


def _trigger_search(platform: str, query: str, limit: int) -> str:
    """Trigger a BrightData dataset collection and return the snapshot_id."""
    dataset_id = BRIGHTDATA_DATASET_IDS[platform]
    payload = {
        "dataset_id": dataset_id,
        "include_errors": False,
        "data": [{"keyword": query, "count": limit}],
    }
    resp = requests.post(
        f"{BRIGHTDATA_BASE_URL}/trigger",
        headers=_brightdata_headers(),
        json=payload,
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["snapshot_id"]


def _poll_snapshot(snapshot_id: str) -> list[dict]:
    """Poll until the snapshot is ready, then return the results."""
    deadline = time.time() + BRIGHTDATA_POLL_TIMEOUT
    while time.time() < deadline:
        resp = requests.get(
            f"{BRIGHTDATA_BASE_URL}/snapshot/{snapshot_id}",
            headers=_brightdata_headers(),
            params={"format": "json"},
            timeout=30,
        )
        if resp.status_code == 200:
            return resp.json()
        if resp.status_code == 202:
            time.sleep(BRIGHTDATA_POLL_INTERVAL)
            continue
        resp.raise_for_status()
    raise TimeoutError(f"BrightData snapshot {snapshot_id} did not complete in time")


def scrape_profiles(platform: str, query: str, limit: int) -> list[dict]:
    """Return raw profile dicts from BrightData."""
    print(f"  Triggering BrightData scrape: platform={platform}, query='{query}', limit={limit}")
    try:
        snapshot_id = _trigger_search(platform, query, limit)
        print(f"  Snapshot triggered ({snapshot_id}), waiting for results...")
        results = _poll_snapshot(snapshot_id)
        print(f"  Got {len(results)} raw profiles")
        return results
    except requests.HTTPError as e:
        if e.response is not None and e.response.status_code == 429:
            print("  Rate limit hit — retrying with smaller batch...")
            time.sleep(10)
            snapshot_id = _trigger_search(platform, query, max(limit // 2, 5))
            return _poll_snapshot(snapshot_id)
        raise


def normalise_profile(raw: dict, platform: str) -> dict:
    """Extract a consistent set of fields regardless of platform response shape."""
    followers = (
        raw.get("followers_count")
        or raw.get("followers")
        or raw.get("subscriber_count")
        or 0
    )
    if isinstance(followers, str):
        followers = int(re.sub(r"[^\d]", "", followers) or 0)

    handle = (
        raw.get("username")
        or raw.get("handle")
        or raw.get("channel_name")
        or raw.get("name", "")
    )
    if handle and not handle.startswith("@"):
        handle = f"@{handle}"

    return {
        "name":      raw.get("full_name") or raw.get("name") or handle,
        "handle":    handle,
        "platform":  platform.capitalize(),
        "followers": followers,
        "bio":       raw.get("biography") or raw.get("bio") or raw.get("description") or "",
        "country":   raw.get("country") or raw.get("location") or "",
        "email":     _extract_email(raw.get("biography") or raw.get("bio") or ""),
    }


def _extract_email(text: str) -> str:
    """Pull the first email address from a bio string, if any."""
    match = re.search(r"[a-zA-Z0-9._%+\-]+@[a-zA-Z0-9.\-]+\.[a-zA-Z]{2,}", text)
    return match.group(0) if match else ""


# ---------------------------------------------------------------------------
# Claude scoring
# ---------------------------------------------------------------------------

SCORE_SYSTEM = ("You are a paid-social creative strategist for a travel brand. You judge whether a "
                "creator would produce short-form video that works as paid ad creative. "
                "Respond with valid JSON only, no extra text.")

# The goal of the gifting programme is to get short-form content Simify can run as
# PAID ADS. So we score for ad-worthiness, not just "is it travel". Bias toward
# personality-led gen-z/millennial lifestyle+travel creators who film themselves.
SCORE_TEMPLATE = """\
Simify (a travel eSIM brand) gifts creators so they make short-form videos Simify can run as PAID ADS.
Judge how good this creator would be for that.

Profile:
  Handle: {handle}
  Platform: {platform}
  Followers: {followers}
  Bio: {bio}
  Country: {country}
  Campaign market (soft preference, not a hard filter): {market}
  Search niche: {niche}

Score 1-10, rewarding:
- Personality-led, on-camera creator (a real face/voice), NOT faceless drone/stock/compilation channels
- Lifestyle + travel content that suits short-form (Shorts/Reels/TikTok-style, snackable, hook-y)
- Would make natural, authentic, brand-safe PAID AD creative (energetic, well-shot, relatable)
- Reads as a gen-z or millennial creator (younger, relatable to travellers)
- English-speaking, reaching a travel-intent Western audience
- Authentic engagement for their size (not inflated), active/posting recently

Penalise: faceless channels, stock-footage or AI-voice compilations, dormant accounts,
non-English audiences unlikely to fit Western paid-ad targeting, pure long-form-only talking heads.

Respond with JSON exactly:
{{"score": <int 1-10>, "reason": "<one sentence: would this make good short-form ad content?>", "recommended": <true|false>}}

recommended = true if score >= 7"""


def score_profile(profile: dict, market: str, niche: str, client: anthropic.Anthropic) -> Optional[dict]:
    """Ask Claude Haiku to score a profile. Returns score dict or None on failure."""
    prompt = SCORE_TEMPLATE.format(
        handle=profile["handle"],
        platform=profile["platform"],
        followers=profile["followers"],
        bio=profile["bio"][:500],
        country=profile["country"],
        market=market,
        niche=niche,
    )
    resp = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=256,
        system=SCORE_SYSTEM,
        messages=[{"role": "user", "content": prompt}],
    )
    raw = resp.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1].lstrip("json").strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError:
        s, e = raw.find("{"), raw.rfind("}") + 1
        if s >= 0 and e > s:
            return json.loads(raw[s:e])
        return None


# ---------------------------------------------------------------------------
# Spreadsheet helpers
# ---------------------------------------------------------------------------

def load_existing_handles(path: Path) -> set[str]:
    """Return the set of dedup keys (URL, else handle, else name) already in the spreadsheet."""
    if not path.exists():
        return set()
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fields = _sheet_fields(ws)
    keys = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        entry = {f: v for f, v in zip(fields, row) if f}
        k = _dedup_key(entry)
        if k:
            keys.add(k)
    return keys


def ensure_spreadsheet(path: Path):
    """Create the spreadsheet with headers if it doesn't exist."""
    if path.exists():
        return
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(SPREADSHEET_COLUMNS)
    wb.save(path)


def append_prospects(path: Path, rows: list[dict]):
    """Append new prospect rows, writing each field into whichever column the sheet has for it."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    fields = _sheet_fields(ws)
    for r in rows:
        r.setdefault("status", "New")
        ws.append([r.get(f, "") if f else "" for f in fields])
    wb.save(path)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def run(
    market: str,
    platform: str,
    niche: str,
    min_followers: int,
    max_followers: int,
    count: int,
):
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("ANTHROPIC_API_KEY is not set in .env")

    client = anthropic.Anthropic(api_key=api_key)
    platform = platform.lower()

    if platform not in BRIGHTDATA_DATASET_IDS:
        raise ValueError(f"Platform must be one of: {', '.join(BRIGHTDATA_DATASET_IDS)}")

    print(f"\nSimify Influencer Discovery")
    print(f"  Market:   {market}")
    print(f"  Platform: {platform.capitalize()}")
    print(f"  Niche:    {niche}")
    print(f"  Followers:{min_followers:,}–{max_followers:,}")
    print(f"  Target:   {count} prospects\n")

    # Step 1 — scrape
    query = f"{niche} {market} travel"
    raw_profiles = scrape_profiles(platform, query, limit=count * 3)

    # Step 2 — normalise + follower filter
    profiles = []
    for raw in raw_profiles:
        p = normalise_profile(raw, platform)
        if min_followers <= p["followers"] <= max_followers:
            profiles.append(p)

    print(f"\n{len(profiles)} profiles within follower range after filtering")

    if not profiles:
        print("Nothing to score. Try a broader niche or follower range.")
        return

    # Step 3 — score with Claude Haiku
    print("Scoring profiles with Claude Haiku...")
    scored = []
    for p in profiles:
        result = score_profile(p, market, niche, client)
        if result is None:
            continue
        if result.get("recommended") or result.get("score", 0) >= 7:
            scored.append({**p, "score": result["score"], "notes": f"Score {result['score']}/10 — {result['reason']}", "niche": niche})
            print(f"  PASS  {p['handle']:<30} score={result['score']}  {result['reason'][:60]}")
        else:
            print(f"  skip  {p['handle']:<30} score={result.get('score', '?')}")

    print(f"\n{len(scored)} prospects passed the filter")

    if not scored:
        print("No prospects passed. Try adjusting niche or market.")
        return

    # Step 4 — deduplicate against spreadsheet
    ensure_spreadsheet(SPREADSHEET_PATH)
    existing = load_existing_handles(SPREADSHEET_PATH)
    new_prospects = [p for p in scored if _dedup_key(p) not in existing]

    dupes = len(scored) - len(new_prospects)
    if dupes:
        print(f"{dupes} already in spreadsheet — skipped")

    if not new_prospects:
        print("All prospects are already in the spreadsheet.")
        return

    # Step 5 — write to spreadsheet
    append_prospects(SPREADSHEET_PATH, new_prospects)

    emails_found = sum(1 for p in new_prospects if p["email"])
    print(f"\nDone.")
    print(f"  {len(new_prospects)} new prospects added to {SPREADSHEET_PATH.name}")
    print(f"  {emails_found} had email addresses in their bio")
    print(f"  Open the spreadsheet to review, then run email_agent.py for outreach drafts")


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Discover influencer prospects for Simify")
    parser.add_argument("--market",       required=True, help="Target market, e.g. 'UK', 'New Zealand'")
    parser.add_argument("--platform",     required=True, choices=["instagram", "tiktok", "youtube"])
    parser.add_argument("--niche",        default="travel", help="Content niche, e.g. 'travel', 'digital nomad'")
    parser.add_argument("--min-followers",type=int, default=20000)
    parser.add_argument("--max-followers",type=int, default=500000)
    parser.add_argument("--count",        type=int, default=20, help="Target number of new prospects")
    args = parser.parse_args()

    run(
        market=args.market,
        platform=args.platform,
        niche=args.niche,
        min_followers=args.min_followers,
        max_followers=args.max_followers,
        count=args.count,
    )

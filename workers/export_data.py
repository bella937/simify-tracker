"""
Builds docs/data/creators.json — the file the dashboard fetches.

Merges THREE private spreadsheets, deduped by normalised YouTube handle:
  1. Simify_YouTube_MicroNano_Prospects.xlsx  "★ Priority Outreach (New)"  (primary reference)
  2. Simify_YouTube_MicroNano_Prospects.xlsx  "Wider Micro-Nano Pool"
  3. Simify_Influencer_Prospects.xlsx          "Prospect List"  (discovery-appended leads)

Deliberately STRIPS PII: contact emails / socials in the source sheets are
NEVER written to creators.json (the published site carries no personal data —
the full contact data stays only in the private xlsx files).

Runnable standalone:  python export_data.py   (from the workers/ folder)
"""

import json
import os
import re
from datetime import datetime, timezone

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
NEW_XLSX = os.path.join(HERE, "Simify_YouTube_MicroNano_Prospects.xlsx")
OLD_XLSX = os.path.join(HERE, "Simify_Influencer_Prospects.xlsx")
OUT = os.path.join(HERE, "..", "docs", "data", "creators.json")

# spreadsheet Status -> [dashboard css class, label]. Labels must match the
# dashboard's group order (docs/index.html ORDER array); classes must match the
# .pillst.* rules (out / repl / prod / live).
STATUS_MAP = {
    "": ["out", "New lead"],
    "new": ["out", "New lead"],
    "new lead": ["out", "New lead"],
    "not contacted": ["out", "New lead"],
    "contacted": ["out", "Outreached"],
    "outreached": ["out", "Outreached"],
    "replied": ["repl", "Replied"],
    "negotiating": ["repl", "Negotiating"],
    "onboarding": ["prod", "Onboarding"],
    "in production": ["prod", "In production"],
    "active": ["live", "Live"],
    "live": ["live", "Live"],
    "declined": ["out", "Declined"],
}
DEFAULT_STATUS = ["out", "New lead"]

# Base/Country -> normalised market code. Known ones map; everything else passes
# through as-is (the "Other" fallback) so codes like IN/JP/CA aren't dropped.
MARKET_MAP = {
    "australia": "AU", "au": "AU", "aus": "AU",
    "united kingdom": "UK", "uk": "UK", "gb": "UK", "england": "UK", "great britain": "UK",
    "new zealand": "NZ", "nz": "NZ",
    "united states": "US", "us": "US", "usa": "US", "united states of america": "US",
    "canada": "CA", "ca": "CA",
    "india": "IN", "in": "IN",
    "japan": "JP", "jp": "JP",
    "ireland": "IE", "ie": "IE",
    "germany": "DE", "de": "DE",
    "france": "FR", "fr": "FR",
    "singapore": "SG", "sg": "SG",
    "philippines": "PH", "ph": "PH",
    "south africa": "ZA", "za": "ZA",
}


def handle_from_url(value):
    """Derive @handle from either a bare '@handle' or a full YouTube URL."""
    if not value:
        return ""
    m = re.search(r"@([A-Za-z0-9_.\-]+)", str(value))
    return "@" + m.group(1) if m else ""


def dedup_key(handle, name):
    """Normalised key for cross-sheet dedup: handle if present, else the name."""
    if handle:
        return handle.lstrip("@").lower()
    return "name:" + str(name or "").strip().lower()


def market(loc):
    """Map a Base/Country string to a market code; strip '(City)' qualifiers."""
    if not loc:
        return "—"
    s = re.sub(r"\(.*?\)", "", str(loc)).strip()  # drop "(Sydney)", "(QLD)", etc.
    return MARKET_MAP.get(s.lower(), s or "—")


def _compact(n):
    n = float(n)
    if n >= 1_000_000:
        v = n / 1_000_000
        return (f"{v:.1f}".rstrip("0").rstrip(".")) + "M"
    if n >= 1_000:
        v = n / 1_000
        return (f"{v:.1f}".rstrip("0").rstrip(".")) + "K"
    return str(int(n))


def fmt_subs(value):
    """Return a consistent compact display string ('72K', '1.5M') for subs."""
    if value is None:
        return ""
    if isinstance(value, (int, float)):
        return _compact(value)
    s = str(value).strip()
    if not s:
        return ""
    compact = s.replace(",", "").replace(" ", "")
    if re.fullmatch(r"\d+(\.\d+)?[KkMm]", compact):
        return compact[:-1] + compact[-1].upper()
    if re.fullmatch(r"\d+", compact):
        return _compact(int(compact))
    return s  # freeform (e.g. a range) — leave as-is


def status_to_st(value):
    return STATUS_MAP.get(str(value or "").strip().lower(), DEFAULT_STATUS)


def _iter_sheet(path, sheet):
    """Yield dict rows keyed by header for a sheet; skips fully blank rows."""
    wb = openpyxl.load_workbook(path, read_only=True)
    if sheet not in wb.sheetnames:
        wb.close()
        return
    ws = wb[sheet]
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    for row in rows:
        if not any(c is not None for c in row):
            continue
        yield {header[i]: (row[i] if i < len(row) else None) for i in range(len(header))}
    wb.close()


def creator(name, handle, loc, subs, niche, status_val):
    """Build one dashboard-schema record (PII intentionally excluded)."""
    return {
        "name": name,
        "handle": handle_from_url(handle) if not str(handle or "").startswith("@") else handle,
        "market": market(loc),
        "subs": fmt_subs(subs),
        "niche": niche or "",
        "deliv": "—",
        "st": status_to_st(status_val),
        "rb": ["new", "new"],
    }


def main():
    creators = []
    seen = set()

    def add(rec):
        key = dedup_key(rec["handle"], rec["name"])
        if key in seen:
            return False
        seen.add(key)
        creators.append(rec)
        return True

    counts = {"priority": 0, "existing": 0, "pool": 0}

    # 1) Priority Outreach (New) — the primary creator reference, richest data.
    for r in _iter_sheet(NEW_XLSX, "★ Priority Outreach (New)"):
        name = r.get("Channel Name")
        if not name:
            continue
        rec = creator(
            name=name,
            handle=r.get("▶ YouTube (click)"),
            loc=r.get("Base"),
            subs=r.get("Subscribers"),
            niche=r.get("Niche / Content"),
            status_val=r.get("Status"),
        )
        if add(rec):
            counts["priority"] += 1

    # 2) Existing Prospect List — discovery-appended leads.
    for r in _iter_sheet(OLD_XLSX, "Prospect List"):
        name = r.get("Channel Name")
        if not name:
            continue
        rec = creator(
            name=name,
            handle=r.get("YouTube URL"),
            loc=r.get("Country"),
            subs=r.get("Subscribers"),
            niche=r.get("Niche"),
            status_val=r.get("Status"),
        )
        if add(rec):
            counts["existing"] += 1

    # 3) Wider Micro-Nano Pool — no Status column, defaults to "New lead".
    for r in _iter_sheet(NEW_XLSX, "Wider Micro-Nano Pool"):
        name = r.get("Channel Name")
        if not name:
            continue
        rec = creator(
            name=name,
            handle=r.get("▶ YouTube (click)"),
            loc=r.get("Country"),
            subs=r.get("Subscribers"),
            niche=r.get("Niche"),
            status_val=None,  # no status -> default New lead
        )
        if add(rec):
            counts["pool"] += 1

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    payload = {
        "updated": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC"),
        "count": len(creators),
        "creators": creators,
    }
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(creators)} creators to {OUT}")
    print(f"  from Priority Outreach: {counts['priority']}, "
          f"Prospect List: {counts['existing']}, Wider Pool: {counts['pool']} "
          f"(after cross-sheet dedup)")


if __name__ == "__main__":
    main()

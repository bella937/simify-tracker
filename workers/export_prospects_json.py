"""
Export Simify_Influencer_Prospects.xlsx to JSON for the Prospects tab
in the Simify Hub talent-manager dashboard (static HTML, no backend —
the dashboard reads whatever this last wrote).

Run after discovery_agent.py / youtube_discovery_agent.py to refresh the dashboard.
"""

import json
from datetime import datetime, timezone
from pathlib import Path

import openpyxl

SPREADSHEET_PATH = Path(__file__).parent / "Simify_Influencer_Prospects.xlsx"
OUTPUT_PATHS = [
    Path.home() / "Downloads" / "Simify Hub" / "data" / "prospects.json",
    Path(__file__).parent / "data" / "prospects.json",
]

# Maps the spreadsheet's actual header text (case/space-insensitive) to the JSON field
# the dashboard reads. Header-driven rather than positional so column reordering in the
# sheet doesn't silently scramble the export.
HEADER_TO_FIELD = {
    "channel name": "name",
    "name": "name",
    "handle": "handle",
    "subscribers": "followers",
    "followers": "followers",
    "platform": "platform",
    "niche": "niche",
    "country": "country",
    "contact email": "email",
    "email": "email",
    "management": "management",
    "youtube url": "url",
    "url": "url",
    "est. rate (usd)": "rate",
    "rate": "rate",
    "status": "status",
    "notes": "notes",
    "last contacted": "last_contacted",
}


def export():
    if not SPREADSHEET_PATH.exists():
        raise FileNotFoundError(f"{SPREADSHEET_PATH} does not exist yet — run a discovery agent first")

    wb = openpyxl.load_workbook(SPREADSHEET_PATH)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    fields = [HEADER_TO_FIELD.get(str(h).strip().lower()) if h else None for h in header_row]
    unmapped = [h for h, f in zip(header_row, fields) if h and f is None]
    if unmapped:
        print(f"Warning: unrecognised columns skipped in export: {unmapped}")

    prospects = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        entry = {f: v for f, v in zip(fields, row) if f}
        entry.setdefault("platform", "YouTube")
        prospects.append(entry)

    payload = json.dumps({
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "prospects": prospects,
    }, indent=2, default=str)

    for output_path in OUTPUT_PATHS:
        try:
            output_path.parent.mkdir(parents=True, exist_ok=True)
            output_path.write_text(payload)
            print(f"Exported {len(prospects)} prospects to {output_path}")
        except OSError as e:
            # Downloads-folder writes can fail under launchd's sandboxing even though
            # the same path works fine from a Terminal session — don't let that block
            # the other destination.
            print(f"Warning: could not write to {output_path}: {e}")


if __name__ == "__main__":
    export()

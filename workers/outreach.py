"""
Creates/updates Simify outreach emails as Gmail DRAFTS (never sends) for sourced
creators who have a direct email. Human reviews and sends from Gmail.
The first line AND the subject line are tailored per creator (free, no AI) from
their name/niche. Idempotent: re-running UPDATES the existing draft to a
recipient rather than duplicating.

SAFETY: only calls drafts().create()/drafts().update(); never messages().send(),
never drafts().send(), never delete. Cannot send or delete mail.
"""
import argparse, base64, html, json, os, re
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
import openpyxl
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

SCOPES = ["https://www.googleapis.com/auth/gmail.readonly",
          "https://www.googleapis.com/auth/gmail.compose"]
HERE = os.path.dirname(os.path.abspath(__file__))
XLSX = os.path.join(HERE, "Simify_YouTube_MicroNano_Prospects.xlsx")
SHEET = "★ Priority Outreach (New)"
JSON_PATH = os.path.join(HERE, "..", "docs", "data", "creators.json")

def first_line(niche):
    if niche:
        seg = re.split(r"[,/(–—|]", str(niche))[0].strip().lower()
        if seg:
            return f"Love your {seg} content on YouTube!"
    return "Love what you're doing on YouTube!"

def greet_name(name):
    if not name: return "there"
    import re as _re
    m=_re.search(r"\(([^)]+)\)", str(name))          # "(Kev)" -> Kev
    if m:
        c=m.group(1).strip().split()
        if c: return c[0]
    s=_re.sub(r"^(hey|hi|hello)\s+","",str(name).strip(),flags=_re.I)  # drop leading greeting
    w=_re.split(r"[\s(\u2013\u2014-]",s)[0].strip()
    STOP={"aussie","real","the","two","my","team","world","adventures","travel","little","big","one","just","not"}
    if not w or w.lower() in STOP or not w[:1].isalpha(): return "there"
    return w

def clean_channel(name):
    """Readable channel label for the subject when there's no usable first name."""
    if not name: return "your channel"
    s = re.sub(r"^(hey|hi|hello)\s+", "", str(name).strip(), flags=re.I)
    s = re.split(r"\s[–—|:]\s", s)[0].strip()   # take text before  - | :
    return s or "your channel"

def subject_line(name):
    """Tailored, unique-per-creator subject (no AI): first name if we have one,
    otherwise the channel name."""
    label = greet_name(name)
    if not label or label == "there":
        label = clean_channel(name)
    return f"{label} × Simify - gifted eSIM + 15% \U0001F381"  # 🎁 (plain hyphen, no em dash)

def body_text(first, niche):
    return (f"Hey {first},\n\n"
            f"{first_line(niche)}\n\n"
            "I'm Bella from Simify - we're a Travel eSIM brand trusted by 1M+ travellers, "
            "and we're looking for creators to join our affiliate programme. Here's how it works:\n\n"
            "\U0001F381 We'll gift you a $100 USD eSIM voucher\n"
            "\U0001F4F1 Share your Simify experience in a YouTube Short or a video integration\n"
            "\U0001F4B8 Earn 15% commission on every sale through your unique discount code\n"
            "\U0001F680 We'll also feature your content in our paid campaigns to boost your reach "
            "and help you grow your audience\n\n"
            "\U0001F449 See the offer + build your content plan (3 mins): https://simify-creators.pages.dev\n\n"
            "Let me know if you're interested and I'll send over all the details!\n\n"
            "Bella\nPartnerships Manager | Simify\nbella@simify.com")

def body_html(first, niche, signature=""):
    fn = html.escape(first)
    fl = html.escape(first_line(niche))
    # Close with Bella's real Gmail signature (branded HTML) when we have it;
    # fall back to a plain text sign-off otherwise. API drafts don't auto-append
    # the Gmail signature, so we embed it here (no risk of it doubling on send).
    signoff = signature or "Bella<br>Partnerships Manager | Simify<br>bella@simify.com"
    return (
        '<div style="font-family:Arial,Helvetica,sans-serif;font-size:14px;'
        'line-height:1.5;color:#111">'
        f"Hey {fn},<br><br>"
        f"{fl}<br><br>"
        '<a href="https://simify-creators.pages.dev" style="text-decoration:none"><img src="https://simify-creators.pages.dev/teaser.png" alt="Simify Creator Programme" width="520" style="width:100%;max-width:520px;border-radius:14px;display:block;margin:2px 0 16px"></a>'
        "I'm Bella from <b>Simify</b> - we're a Travel eSIM brand trusted by 1M+ travellers, "
        "and we're looking for creators to join our <b>affiliate programme</b>. "
        "Here's how it works:<br><br>"
        "\U0001F381 We'll gift you a <b>$100 USD eSIM voucher</b><br>"
        "\U0001F4F1 Share your Simify experience in a YouTube Short or a video integration<br>"
        "\U0001F4B8 Earn <b>15% commission</b> on every sale through your unique discount code<br>"
        "\U0001F680 We'll also feature your content in our paid campaigns to boost your reach "
        "and help you grow your audience<br><br>"
        "\U0001F449 <b>See the offer + build your content plan in 3 minutes:</b> "
        '<a href="https://simify-creators.pages.dev">simify-creators.pages.dev</a><br><br>'
        "Let me know if you're interested and I'll send over all the details!<br><br>"
        f"{signoff}"
        "</div>")

def get_signature(svc):
    """Return Bella's default Gmail sign-as signature (HTML), or '' if unavailable."""
    try:
        res = svc.users().settings().sendAs().list(userId="me").execute()
        sends = res.get("sendAs", [])
        for sa in sends:
            if sa.get("isDefault") and sa.get("signature"):
                return sa["signature"]
        for sa in sends:
            if sa.get("signature"):
                return sa["signature"]
    except Exception as e:
        print(f"  [warn] couldn't read Gmail signature ({type(e).__name__}); using text sign-off")
    return ""

def gmail():
    c = Credentials.from_authorized_user_file("token.json", SCOPES)
    if not c.valid and c.expired and c.refresh_token:
        c.refresh(Request()); open("token.json","w").write(c.to_json())
    return build("gmail", "v1", credentials=c)

def existing_drafts_by_recipient(svc):
    m = {}
    res = svc.users().drafts().list(userId="me", maxResults=200).execute()
    for d in res.get("drafts", []):
        meta = svc.users().drafts().get(userId="me", id=d["id"], format="metadata").execute()
        hdrs = meta.get("message", {}).get("payload", {}).get("headers", [])
        to = next((h["value"] for h in hdrs if h["name"].lower() == "to"), "")
        if to: m[to.strip().lower()] = d["id"]
    return m

def raw(to, subject, text, html_body):
    # multipart/alternative: plain-text fallback + HTML (so **bold** renders).
    msg = MIMEMultipart("alternative")
    msg["to"] = to; msg["subject"] = subject
    msg.attach(MIMEText(text, "plain", "utf-8"))       # utf-8 so emoji encode correctly
    msg.attach(MIMEText(html_body, "html", "utf-8"))
    return base64.urlsafe_b64encode(msg.as_bytes()).decode()

def main():
    wb = openpyxl.load_workbook(XLSX); ws = wb[SHEET]
    hdr = [c.value for c in ws[1]]; ci = {h: i for i, h in enumerate(hdr)}
    svc = gmail()
    sig = get_signature(svc)
    drafts = existing_drafts_by_recipient(svc)
    n = 0
    for row in ws.iter_rows(min_row=2):
        vals = [c.value for c in row]
        if not any(vals): continue
        email = vals[ci.get("Email")]
        if not email or "@" not in str(email): continue
        name = vals[ci.get("Channel Name")]
        niche = vals[ci.get("Niche / Content")]
        first = greet_name(name)
        to = str(email).strip()
        subject = subject_line(name)
        r = raw(to, subject, body_text(first, niche), body_html(first, niche, sig))
        did = drafts.get(to.lower())
        if did:
            svc.users().drafts().update(userId="me", id=did, body={"message": {"raw": r}}).execute()
            action = "updated"
        else:
            svc.users().drafts().create(userId="me", body={"message": {"raw": r}}).execute()
            action = "created"
        row[ci["Status"]].value = "Contacted"
        print(f"  {action}: {name} <{to}>")
        n += 1
    wb.save(XLSX)
    print(f"Done. {n} draft(s) synced in Gmail with your template (review + send there). No emails sent.")

def draft_from_json(markets=None):
    """Draft to EVERY creator in creators.json that has an email, optionally
    filtered to a set of market codes (e.g. {'AU','UK','NZ'}). Reuses the same
    template + idempotent upsert. Broader than main(), which only reads the
    priority xlsx sheet."""
    creators = json.load(open(JSON_PATH, encoding="utf-8"))["creators"]
    want = {m.strip().upper() for m in markets} if markets else None
    svc = gmail()
    sig = get_signature(svc)
    drafts = existing_drafts_by_recipient(svc)
    n = 0
    for c in creators:
        email = (c.get("email") or "").strip()
        if "@" not in email:
            continue
        if want is not None and str(c.get("market", "")).upper() not in want:
            continue
        name, niche = c.get("name"), c.get("niche")
        first = greet_name(name)
        subject = subject_line(name)
        r = raw(email, subject, body_text(first, niche), body_html(first, niche, sig))
        did = drafts.get(email.lower())
        if did:
            svc.users().drafts().update(userId="me", id=did, body={"message": {"raw": r}}).execute()
            action = "updated"
        else:
            svc.users().drafts().create(userId="me", body={"message": {"raw": r}}).execute()
            action = "created"
        print(f"  {action}: {name} <{email}> [{c.get('market')}]")
        n += 1
    print(f"Done. {n} draft(s) synced in Gmail (review + send there). No emails sent.")

if __name__ == "__main__":
    ap = argparse.ArgumentParser(description="Create Simify outreach Gmail DRAFTS (never sends).")
    ap.add_argument("--from-json", action="store_true",
                    help="draft to all emailable creators in creators.json (not just the priority sheet)")
    ap.add_argument("--markets", default=None,
                    help="comma-separated market filter for --from-json, e.g. AU,UK,NZ")
    args = ap.parse_args()
    if args.from_json:
        draft_from_json(args.markets.split(",") if args.markets else None)
    else:
        main()
